# TAGs Transporte Furlong - Version unificada. Flask + SQLite + pdfplumber. Sin IA, sin API keys.
import csv, io, os, re, sqlite3, base64, json
from datetime import datetime
from flask import Flask, g, request, redirect, session, render_template, send_file, flash, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import load_workbook, Workbook

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "cambiar-en-produccion")
DB = os.environ.get("DB_PATH", "furlong.db")

# ================= BASE DE DATOS =================
def db():
    if "db" not in g:
        g.db = sqlite3.connect(DB)
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(_):
    d = g.pop("db", None)
    if d: d.close()

def init_db():
    con = sqlite3.connect(DB)
    con.executescript("""
    CREATE TABLE IF NOT EXISTS usuarios(id INTEGER PRIMARY KEY, email TEXT UNIQUE,
        clave TEXT, nombre TEXT, rol TEXT DEFAULT 'operador');
    CREATE TABLE IF NOT EXISTS flota(id INTEGER PRIMARY KEY, tag TEXT, patente TEXT,
        dueno TEXT, equipo TEXT, estado TEXT, marca TEXT);
    CREATE TABLE IF NOT EXISTS movimientos(id INTEGER PRIMARY KEY, fecha TEXT, tag TEXT,
        patente TEXT, concepto TEXT, monto REAL, dueno TEXT, equipo TEXT,
        verificado INTEGER DEFAULT 0, archivo TEXT, creado TEXT);
    CREATE TABLE IF NOT EXISTS archivos(nombre TEXT PRIMARY KEY, tipo TEXT,
        total_doc REAL DEFAULT 0, contenido TEXT, mime TEXT, creado TEXT);
    CREATE TABLE IF NOT EXISTS resumenes(id INTEGER PRIMARY KEY, archivo TEXT, banco TEXT,
        titular TEXT, periodo TEXT, vencimiento TEXT, total_resumen REAL, creado TEXT);
    CREATE TABLE IF NOT EXISTS gastos(id INTEGER PRIMARY KEY, resumen_id INTEGER,
        fecha TEXT, concepto TEXT, categoria TEXT, monto REAL);
    CREATE INDEX IF NOT EXISTS idx_mov_tag ON movimientos(tag);
    CREATE INDEX IF NOT EXISTS idx_flota_tag ON flota(tag);
    """)
    if not con.execute("SELECT 1 FROM usuarios").fetchone():
        con.execute("INSERT INTO usuarios(email,clave,nombre,rol) VALUES(?,?,?,?)",
                    ("admin@transportefurlong.com.ar", generate_password_hash("admin"),
                     "Admin Furlong", "admin"))
    con.commit(); con.close()

# ================= UTILIDADES =================
def norm_tag(v):
    s = re.sub(r"[^A-Z0-9]", "", str(v or "").upper())
    d = re.sub(r"[^0-9]", "", s)
    return d[-8:] if len(d) >= 8 else d

def norm_monto(v):
    if isinstance(v, (int, float)): return float(v)
    s = re.sub(r"[^\d.,\-]", "", str(v or ""))
    if not s: return 0.0
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".") if s.rfind(",") > s.rfind(".") else s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")
    try: return float(s)
    except ValueError: return 0.0

RE_FECHA = re.compile(r"\b(\d{1,2}[/.\-]\d{1,2}[/.\-]\d{2,4})\b")

def fmt_fecha(v):
    if isinstance(v, datetime): return v.strftime("%d/%m/%Y")
    m = RE_FECHA.search(str(v or ""))
    if not m: return str(v or "").strip()[:10]
    d, mo, y = re.split(r"[/.\-]", m.group(1))
    if len(y) == 2: y = "20" + y
    try: return f"{int(d):02d}/{int(mo):02d}/{int(y):04d}"
    except ValueError: return m.group(1)

def pesos(n):
    return f"${n:,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")

app.jinja_env.filters["pesos"] = pesos

KEYS = {
    "tag": ["NUMERO DE TAG", "NRO TAG", "N TAG", "TAG", "DISPOSITIVO"],
    "patente": ["PATENTE", "DOMINIO"],
    "dueno": ["RESPONSABLE DE USUARIO", "RESPONSABLE", "DUENO", "DUEÑO", "TITULAR"],
    "equipo": ["EQUIPOS", "EQUIPO"], "estado": ["ESTADO"], "marca": ["MARCA"],
    "fecha": ["FECHA"], "concepto": ["CONCEPTO", "DESCRIPCION", "ESTACION", "DETALLE", "PEAJE"],
    "monto": ["IMPORTE", "MONTO", "VALOR", "TOTAL"],
}

def leer_filas(archivo):
    nombre = archivo.filename.lower()
    if nombre.endswith(".csv"):
        texto = archivo.read().decode("utf-8-sig", errors="replace")
        sep = ";" if texto.count(";") > texto.count(",") else ","
        return [row for row in csv.reader(io.StringIO(texto), delimiter=sep)]
    wb = load_workbook(archivo, read_only=True, data_only=True)
    filas = []
    for hoja in wb.worksheets:
        for row in hoja.iter_rows(values_only=True):
            filas.append(list(row))
    return filas

def mapear_columnas(filas, campos):
    for i, fila in enumerate(filas[:15]):
        celdas = [re.sub(r"[^A-ZÑ ]", "", str(c or "").upper().strip()) for c in fila]
        mapa = {}
        for campo in campos:
            for j, celda in enumerate(celdas):
                if celda and any(k in celda for k in KEYS[campo]) and j not in mapa.values():
                    mapa[campo] = j; break
        if len(mapa) >= 2:
            return mapa, i + 1
    return None, 0

# ================= EXTRACCION PDF (sin IA) =================
RE_MONTO = re.compile(r"\$?\s*\d{1,3}(?:\.\d{3})*,\d{2}|\$?\s*\d+,\d{2}")
RE_TAG = re.compile(r"\b[A-Z]{0,4}\d{8,14}\b")
RE_TOTAL = re.compile(r"(TOTAL\s+GRAVADO|SUBTOTAL|IMPORTE\s+TOTAL|TOTAL)\D*?([\d.,]+,\d{2})", re.I)
RE_TOTAL_TARJETA = re.compile(r"(TOTAL\s+A\s+PAGAR|SALDO\s+ACTUAL|TOTAL\s+CONSUMOS|PAGO\s+TOTAL)\D*?([\d.,]+,\d{2})", re.I)
RE_PERIODO = re.compile(r"(CIERRE|PERIODO|RESUMEN)\D{0,20}(\d{1,2}[/.\-]\d{1,2}[/.\-]\d{2,4})", re.I)
RE_VENC = re.compile(r"(VENCIMIENTO|VENCE|VTO)\D{0,20}(\d{1,2}[/.\-]\d{1,2}[/.\-]\d{2,4})", re.I)
BANCOS = ["GALICIA", "SANTANDER", "BBVA", "MACRO", "NACION", "PROVINCIA", "ICBC", "HSBC",
          "PATAGONIA", "SUPERVIELLE", "CREDICOOP", "COMAFI", "ITAU", "CITI", "BIND", "HIPOTECARIO",
          "CIUDAD", "SANTA FE", "CORDOBA", "ENTRE RIOS", "SAN JUAN", "ROELA", "COLUMBIA",
          "BRUBANK", "UALA", "UALÁ", "MERCADO PAGO", "MERCADOPAGO", "REBA", "WILOBANK",
          "VISA", "MASTERCARD", "MAESTRO", "AMEX", "AMERICAN EXPRESS", "CABAL", "NARANJA", "CENCOSUD"]

def texto_pdf(archivo):
    import pdfplumber
    texto = ""
    with pdfplumber.open(archivo) as pdf:
        for pag in pdf.pages:
            texto += (pag.extract_text() or "") + "\n"
    if not texto.strip():
        raise ValueError("el PDF no tiene texto legible (puede ser un escaneo).")
    return texto

def categorizar(desc):
    d = desc.upper()
    if any(k in d for k in ("PEAJE", "AUSA", "AUSOL", "AUTOPISTA", "TELEPASE", "TELEPEAJE", "AUBASA",
        "CORREDOR VIAL", "RICCHERI", "ILLIA", "DELLEPIANE", "ACCESO OESTE", "ACCESO NORTE",
        "CAMINOS DEL", "AUTOPISTAS DEL", "RUTAS DEL", "PANAMERICANA", "PASE URBANO")): return "PEAJE"
    if any(k in d for k in ("YPF", "AXION", "SHELL", "PUMA", "GNC", "COMBUSTIBLE", "NAFTA", "GASOIL",
        "REFINOR", "ESTACION DE SERVICIO", "LUBRICANTES YPF")): return "COMBUSTIBLE"
    if any(k in d for k in ("SERVICE", "TALLER", "REPUESTO", "GOMERIA", "NEUMATICO", "LUBRICANTE",
        "MECANIC", "FILTRO", "ACEITE", "AUTOELEVADOR", "CUBIERTA")): return "MANTENIMIENTO"
    return "GASTO"

def parsear_pdf_peajes(archivo):
    texto = texto_pdf(archivo)
    total_doc, prioridad = 0.0, 9
    orden = {"TOTAL GRAVADO": 0, "SUBTOTAL": 1, "IMPORTE TOTAL": 2, "TOTAL": 3}
    for m in RE_TOTAL.finditer(texto):
        clave = re.sub(r"\s+", " ", m.group(1).upper())
        p = orden.get(clave, 3); v = norm_monto(m.group(2))
        if v > 0 and (p < prioridad or (p == prioridad and v > total_doc)):
            total_doc, prioridad = v, p
    m = RE_FECHA.search(texto)
    fecha_doc = fmt_fecha(m.group(1)) if m else ""
    items = []
    for linea in texto.split("\n"):
        lu = linea.upper()
        if any(p in lu for p in ("TOTAL", "SUBTOTAL", "IVA", "CUIT", "C.U.I.T", "SALDO", "VENCIMIENTO")):
            continue
        tags = RE_TAG.findall(lu); montos = RE_MONTO.findall(linea)
        if not tags or not montos: continue
        tag = norm_tag(tags[0])
        if len(tag) < 8: continue
        monto = norm_monto(montos[-1])
        if monto <= 0: continue
        f = RE_FECHA.search(linea)
        concepto = re.sub(r"\s{2,}", " ", RE_FECHA.sub("", RE_MONTO.sub("", RE_TAG.sub("", linea)))).strip(" -|$") or "Peaje"
        items.append({"fecha": fmt_fecha(f.group(1)) if f else fecha_doc, "tag": tag,
                      "concepto": concepto[:80], "monto": monto})
    return items, total_doc

def parsear_pdf_tarjeta(archivo):
    texto = texto_pdf(archivo)
    tu = texto.upper()
    banco = next((b for b in BANCOS if b in tu), "Banco")
    titular = ""
    for pat in (r"TITULAR\W{0,5}([A-ZÑÁÉÍÓÚ][A-ZÑÁÉÍÓÚ ,.]{3,40})",
                r"SR[A]?\.?\W{1,3}([A-ZÑÁÉÍÓÚ][A-ZÑÁÉÍÓÚ ,.]{3,40})"):
        m = re.search(pat, tu)
        if m:
            titular = re.sub(r"\s{2,}", " ", m.group(1)).strip(" ,.").title()
            break
    m = RE_PERIODO.search(texto)
    periodo = fmt_fecha(m.group(2)) if m else (fmt_fecha(RE_FECHA.search(texto).group(1)) if RE_FECHA.search(texto) else "")
    m = RE_VENC.search(texto)
    venc = fmt_fecha(m.group(2)) if m else "-"
    total = 0.0
    for m in RE_TOTAL_TARJETA.finditer(texto):
        total = max(total, norm_monto(m.group(2)))
    items = []
    for linea in texto.split("\n"):
        lu = linea.upper()
        if any(p in lu for p in ("TOTAL", "SALDO", "PAGO MINIMO", "CIERRE", "VENCIMIENTO", "CUIT", "LIMITE", "TASA", "INTERES")):
            continue
        f = RE_FECHA.search(linea); montos = RE_MONTO.findall(linea)
        if not f or not montos: continue
        monto = norm_monto(montos[-1])
        if monto <= 0: continue
        desc = re.sub(r"\s{2,}", " ", RE_MONTO.sub("", RE_FECHA.sub("", linea))).strip(" -|$*")
        if len(desc) < 3: continue
        items.append({"fecha": fmt_fecha(f.group(1)), "concepto": desc[:80],
                      "categoria": categorizar(desc), "monto": monto})
    return {"banco": banco.title(), "titular": titular, "periodo": periodo,
            "vencimiento": venc, "total": total}, items

def extraer_items_peajes(arch, contenido):
    if arch.filename.lower().endswith(".pdf"):
        return parsear_pdf_peajes(io.BytesIO(contenido))
    arch.seek(0)
    filas = leer_filas(arch)
    mapa, inicio = mapear_columnas(filas, ["fecha", "tag", "patente", "concepto", "monto"])
    if not mapa or "monto" not in mapa:
        raise ValueError("no se detectaron las columnas (se necesita al menos IMPORTE/MONTO).")
    items = []
    for fila in filas[inicio:]:
        get = lambda c: fila[mapa[c]] if c in mapa and mapa[c] < len(fila) else ""
        monto = norm_monto(get("monto"))
        if monto == 0: continue
        items.append({"fecha": fmt_fecha(get("fecha")), "tag": norm_tag(get("tag")),
                      "patente": str(get("patente") or "").strip().upper(),
                      "concepto": str(get("concepto") or "Peaje").strip(), "monto": monto})
    return items, 0.0

def guardar_archivo(nombre, tipo, total_doc, contenido, mime):
    b64 = base64.b64encode(contenido).decode() if contenido else ""
    db().execute("""INSERT OR REPLACE INTO archivos(nombre,tipo,total_doc,contenido,mime,creado)
        VALUES(?,?,?,?,?,?)""", (nombre, tipo, total_doc, b64, mime,
        datetime.now().strftime("%d/%m/%Y %H:%M")))

# ================= AUTENTICACION =================
@app.before_request
def requiere_login():
    if request.endpoint not in ("login", "static") and not session.get("uid"):
        return redirect("/login")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        u = db().execute("SELECT * FROM usuarios WHERE email=?",
                         (request.form["email"].strip().lower(),)).fetchone()
        if u and check_password_hash(u["clave"], request.form["clave"]):
            session.update(uid=u["id"], email=u["email"], nombre=u["nombre"], rol=u["rol"])
            return redirect("/")
        flash("Credenciales invalidas.")
    return render_template("login.html")

@app.route("/salir")
def salir():
    session.clear(); return redirect("/login")

# ================= 1) TABLERO =================
@app.route("/")
def dashboard():
    d = db()
    kpi = d.execute("""SELECT COUNT(*) n, COALESCE(SUM(monto),0) total,
        COUNT(DISTINCT NULLIF(patente,'')) patentes, COUNT(DISTINCT NULLIF(dueno,'')) duenos,
        COALESCE(SUM(verificado),0) ok FROM movimientos""").fetchone()
    top = d.execute("""SELECT COALESCE(NULLIF(dueno,''),'Sin asignar') dueno, SUM(monto) t,
        COUNT(*) n FROM movimientos GROUP BY 1 ORDER BY t DESC LIMIT 10""").fetchall()
    dist = d.execute("""SELECT CASE WHEN verificado=1 THEN 'Verificados en flota'
        ELSE 'Sin match de TAG' END k, COUNT(*) n, SUM(monto) t FROM movimientos GROUP BY 1""").fetchall()
    ult = d.execute("SELECT * FROM movimientos ORDER BY id DESC LIMIT 20").fetchall()
    maximo = max([r["t"] for r in top], default=1)
    return render_template("dashboard.html", kpi=kpi, top=top, dist=dist, ult=ult, maximo=maximo)

# --- Endpoint JSON para modales de doble-click (drill-down) ---
@app.route("/detalle/<tipo>")
def detalle(tipo):
    d = db()
    arg = request.args.get("v", "")
    if tipo == "monto" or tipo == "operaciones":
        rows = d.execute("SELECT * FROM movimientos ORDER BY monto DESC LIMIT 1000").fetchall()
        titulo = "Monto total" if tipo == "monto" else "Operaciones"
    elif tipo == "camiones":
        rows = d.execute("""SELECT * FROM movimientos WHERE patente!='' GROUP BY patente
            ORDER BY SUM(monto) DESC""").fetchall()
        titulo = "Camiones activos"
    elif tipo == "duenos":
        rows = d.execute("""SELECT dueno, COUNT(*) n, SUM(monto) monto,
            COUNT(DISTINCT patente) patentes FROM movimientos WHERE dueno!=''
            GROUP BY dueno ORDER BY monto DESC""").fetchall()
        return jsonify(titulo="Responsables", modo="grupo",
                       filas=[dict(r) for r in rows])
    elif tipo == "dueno":
        rows = d.execute("SELECT * FROM movimientos WHERE dueno=? ORDER BY monto DESC", (arg,)).fetchall()
        titulo = f"Detalle: {arg}"
    else:
        rows = d.execute("SELECT * FROM movimientos ORDER BY id DESC LIMIT 500").fetchall()
        titulo = "Movimientos"
    return jsonify(titulo=titulo, modo="detalle", filas=[dict(r) for r in rows])

@app.route("/movimientos")
def movimientos():
    q = request.args.get("q", "").strip()
    dueno = request.args.get("dueno", "").strip()
    vista = request.args.get("vista", "detalle")
    d = db()
    if vista == "vehiculo":
        grupos = d.execute("""SELECT COALESCE(NULLIF(patente,''),'S/D') g, COUNT(*) n,
            SUM(monto) t, MAX(dueno) sub FROM movimientos GROUP BY 1 ORDER BY t DESC""").fetchall()
        return render_template("movimientos.html", vista=vista, grupos=grupos, q=q, movs=[], etiqueta="Vehiculo")
    if vista == "responsable":
        grupos = d.execute("""SELECT COALESCE(NULLIF(dueno,''),'Sin asignar') g, COUNT(*) n,
            SUM(monto) t, COUNT(DISTINCT patente) sub FROM movimientos GROUP BY 1 ORDER BY t DESC""").fetchall()
        return render_template("movimientos.html", vista=vista, grupos=grupos, q=q, movs=[], etiqueta="Responsable")
    filtro = request.args.get("filtro", "")
    args, conds = [], []
    if q:
        conds.append("(tag LIKE ? OR patente LIKE ? OR dueno LIKE ? OR concepto LIKE ?)")
        args += [f"%{q}%"] * 4
    if dueno:
        conds.append("COALESCE(NULLIF(dueno,''),'Sin asignar')=?"); args.append(dueno)
    if filtro == "verificado":
        conds.append("verificado=1")
    elif filtro == "sinmatch":
        conds.append("verificado=0")
    where = (" WHERE " + " AND ".join(conds)) if conds else ""
    movs = d.execute("SELECT * FROM movimientos" + where + " ORDER BY id DESC LIMIT 500", args).fetchall()
    resumen = d.execute("SELECT COUNT(*) n, COALESCE(SUM(monto),0) t FROM movimientos" + where, args).fetchone()
    return render_template("movimientos.html", vista="detalle", movs=movs, q=q or dueno,
                           qtext=q, dueno=dueno, filtro=filtro, resumen=resumen, grupos=[])

@app.route("/movimientos/borrar", methods=["POST"])
def borrar_movimientos():
    ids = request.form.getlist("ids")
    arch = request.form.get("archivo")
    d = db()
    if ids:
        d.executemany("DELETE FROM movimientos WHERE id=?", [(i,) for i in ids])
    elif arch:
        d.execute("DELETE FROM movimientos WHERE archivo=?", (arch,))
    d.commit(); flash("Registros eliminados.")
    return redirect(request.referrer or "/movimientos")

# ================= 2) IMPORTAR =================
@app.route("/importar", methods=["GET", "POST"])
def importar():
    d = db()
    if request.method == "POST" and request.files.get("archivo"):
        arch = request.files["archivo"]
        contenido = arch.read()
        items, total_doc, error = [], 0.0, None
        try:
            items, total_doc = extraer_items_peajes(arch, contenido)
        except Exception as e:
            error = str(e)
        if error:
            flash(f"Error al procesar: {error}")
        elif not items:
            flash("No se encontraron movimientos en el archivo.")
        else:
            flota_map = {f["tag"]: f for f in d.execute("SELECT * FROM flota")}
            nuevos, total = [], 0.0
            for it in items:
                m = flota_map.get(it["tag"])
                nuevos.append((it["fecha"], it["tag"], (m["patente"] if m else it.get("patente", "")),
                               it["concepto"], it["monto"], (m["dueno"] if m else ""),
                               (m["equipo"] if m else ""), 1 if m else 0, arch.filename, datetime.now().isoformat()))
                total += it["monto"]
            d.executemany("""INSERT INTO movimientos(fecha,tag,patente,concepto,monto,
                dueno,equipo,verificado,archivo,creado) VALUES(?,?,?,?,?,?,?,?,?,?)""", nuevos)
            mime = "application/pdf" if arch.filename.lower().endswith(".pdf") else "application/vnd.ms-excel"
            guardar_archivo(arch.filename, "Peajes", total_doc, contenido, mime)
            d.commit()
            ok = sum(1 for n in nuevos if n[7])
            msg = f"Importados {len(nuevos)} movimientos ({pesos(total)}). Verificados: {ok}."
            if total_doc:
                dif = abs(total_doc - total)
                msg += (f" ✔ Control OK ({pesos(total_doc)})." if dif < 1
                        else f" ⚠ Documento declara {pesos(total_doc)}, diferencia {pesos(dif)}.")
            flash(msg)
        return redirect("/importar")
    flota_n = d.execute("SELECT COUNT(*) n FROM flota").fetchone()["n"]
    archivos = d.execute("""SELECT m.archivo, COUNT(*) n, SUM(m.monto) t, COALESCE(a.total_doc,0) td
        FROM movimientos m LEFT JOIN archivos a ON a.nombre=m.archivo
        GROUP BY m.archivo ORDER BY MAX(m.id) DESC""").fetchall()
    return render_template("importar.html", archivos=archivos, flota_n=flota_n)

@app.route("/flota", methods=["GET", "POST"])
def flota():
    d = db()
    if request.method == "POST" and request.files.get("archivo"):
        try:
            filas = leer_filas(request.files["archivo"])
            mapa, inicio = mapear_columnas(filas, ["tag", "patente", "dueno", "equipo", "estado", "marca"])
        except Exception as e:
            mapa, inicio, filas = None, 0, []
            flash(f"Error al leer: {e}")
        if filas and (not mapa or "tag" not in mapa):
            flash("No se detectaron las columnas (se necesita al menos TAG).")
        elif mapa:
            vistos, nuevos = set(), []
            for fila in filas[inicio:]:
                get = lambda c: str(fila[mapa[c]] or "").strip() if c in mapa and mapa[c] < len(fila) else ""
                t = norm_tag(get("tag"))
                if t and t not in vistos:
                    vistos.add(t)
                    nuevos.append((t, get("patente").upper(), get("dueno"), get("equipo"), get("estado"), get("marca")))
            d.execute("DELETE FROM flota")
            d.executemany("INSERT INTO flota(tag,patente,dueno,equipo,estado,marca) VALUES(?,?,?,?,?,?)", nuevos)
            d.commit()
            flash(f"Base actualizada: {len(nuevos)} unidades unicas.")
        return redirect("/flota")
    unidades = d.execute("SELECT * FROM flota ORDER BY dueno, patente").fetchall()
    return render_template("flota.html", unidades=unidades)

# ================= 3) GASTOS TARJETAS =================
@app.route("/gastos", methods=["GET", "POST"])
def gastos():
    d = db()
    if request.method == "POST" and request.files.get("archivo"):
        arch = request.files["archivo"]
        contenido = arch.read()
        try:
            if arch.filename.lower().endswith(".pdf"):
                meta, items = parsear_pdf_tarjeta(io.BytesIO(contenido))
            else:
                arch.seek(0); filas = leer_filas(arch)
                mapa, inicio = mapear_columnas(filas, ["fecha", "concepto", "monto"])
                if not mapa or "monto" not in mapa:
                    raise ValueError("no se detectaron columnas FECHA/DESCRIPCION/MONTO.")
                items = []
                for fila in filas[inicio:]:
                    get = lambda c: fila[mapa[c]] if c in mapa and mapa[c] < len(fila) else ""
                    monto = norm_monto(get("monto"))
                    if monto == 0: continue
                    desc = str(get("concepto") or "").strip()
                    items.append({"fecha": fmt_fecha(get("fecha")), "concepto": desc,
                                  "categoria": categorizar(desc), "monto": monto})
                meta = {"banco": "Excel", "titular": "", "periodo": "", "vencimiento": "-",
                        "total": sum(i["monto"] for i in items)}
            if not items:
                raise ValueError("no se encontraron consumos en el resumen.")
            cur = d.execute("""INSERT INTO resumenes(archivo,banco,titular,periodo,vencimiento,total_resumen,creado)
                VALUES(?,?,?,?,?,?,?)""", (arch.filename, meta["banco"], meta["titular"], meta["periodo"],
                meta["vencimiento"], meta["total"], datetime.now().isoformat()))
            rid = cur.lastrowid
            d.executemany("INSERT INTO gastos(resumen_id,fecha,concepto,categoria,monto) VALUES(?,?,?,?,?)",
                          [(rid, i["fecha"], i["concepto"], i["categoria"], i["monto"]) for i in items])
            mime = "application/pdf" if arch.filename.lower().endswith(".pdf") else "application/vnd.ms-excel"
            guardar_archivo(arch.filename, "Resumen tarjeta", meta["total"], contenido, mime)
            d.commit()
            flash(f"Resumen {meta['banco']} procesado: {len(items)} consumos, total {pesos(meta['total'])}.")
        except Exception as e:
            flash(f"Error al procesar: {e}")
        return redirect("/gastos")
    resumenes = d.execute("""SELECT r.*, COUNT(g.id) n, COALESCE(SUM(g.monto),0) suma,
        COALESCE(SUM(CASE WHEN g.categoria='PEAJE' THEN g.monto ELSE 0 END),0) peajes
        FROM resumenes r LEFT JOIN gastos g ON g.resumen_id=r.id GROUP BY r.id ORDER BY r.id DESC""").fetchall()
    cat = request.args.get("cat", "")
    sql = "SELECT g.*, r.banco, r.archivo FROM gastos g JOIN resumenes r ON r.id=g.resumen_id"
    args = []
    if cat: sql += " WHERE g.categoria=?"; args.append(cat)
    detalle = d.execute(sql + " ORDER BY g.id DESC LIMIT 500", args).fetchall()
    deuda = d.execute("SELECT COALESCE(SUM(total_resumen),0) t, COUNT(*) n FROM resumenes").fetchone()
    cats = d.execute("SELECT categoria, COUNT(*) n, SUM(monto) t FROM gastos GROUP BY categoria ORDER BY t DESC").fetchall()
    return render_template("gastos.html", resumenes=resumenes, detalle=detalle, deuda=deuda, cat=cat, cats=cats)

@app.route("/gastos/borrar/<int:rid>", methods=["POST"])
def borrar_resumen(rid):
    d = db()
    d.execute("DELETE FROM gastos WHERE resumen_id=?", (rid,))
    d.execute("DELETE FROM resumenes WHERE id=?", (rid,))
    d.commit(); flash("Resumen eliminado.")
    return redirect("/gastos")

@app.route("/gastos/exportar")
def exportar_gastos():
    wb = Workbook(); ws = wb.active; ws.title = "Gastos"
    ws.append(["FECHA", "CONCEPTO", "CATEGORIA", "MONTO", "BANCO", "RESUMEN ORIGEN"])
    for m in db().execute("""SELECT g.*, r.banco, r.archivo FROM gastos g
        JOIN resumenes r ON r.id=g.resumen_id ORDER BY r.id, g.id"""):
        ws.append([m["fecha"], m["concepto"], m["categoria"], m["monto"], m["banco"], m["archivo"]])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"gastos_tarjetas_{datetime.now():%Y%m%d}.xlsx")

# ================= 4) REPORTES =================
@app.route("/reportes")
def reportes():
    d = db()
    n = d.execute("SELECT COUNT(*) n FROM movimientos").fetchone()["n"]
    ng = d.execute("SELECT COUNT(*) n FROM gastos").fetchone()["n"]
    historial = d.execute("SELECT nombre,tipo,total_doc,creado FROM archivos ORDER BY rowid DESC").fetchall()
    return render_template("reportes.html", n=n, ng=ng, historial=historial)

@app.route("/reportes/descargar/<path:nombre>")
def descargar_archivo(nombre):
    r = db().execute("SELECT contenido,mime,nombre FROM archivos WHERE nombre=?", (nombre,)).fetchone()
    if not r or not r["contenido"]:
        flash("Archivo no disponible."); return redirect("/reportes")
    data = base64.b64decode(r["contenido"])
    return send_file(io.BytesIO(data), as_attachment=True, download_name=r["nombre"],
                     mimetype=r["mime"] or "application/octet-stream")

@app.route("/reportes/borrar", methods=["POST"])
def borrar_archivo():
    nombres = request.form.getlist("nombres") or ([request.form["nombre"]] if request.form.get("nombre") else [])
    d = db()
    for nombre in nombres:
        d.execute("DELETE FROM movimientos WHERE archivo=?", (nombre,))
        for r in d.execute("SELECT id FROM resumenes WHERE archivo=?", (nombre,)).fetchall():
            d.execute("DELETE FROM gastos WHERE resumen_id=?", (r["id"],))
        d.execute("DELETE FROM resumenes WHERE archivo=?", (nombre,))
        d.execute("DELETE FROM archivos WHERE nombre=?", (nombre,))
    d.commit(); flash(f"{len(nombres)} archivo(s) y sus registros eliminados.")
    return redirect("/reportes")

@app.route("/exportar")
def exportar():
    wb = Workbook(); ws = wb.active; ws.title = "Movimientos"
    ws.append(["FECHA", "TAG", "PATENTE", "CONCEPTO", "MONTO", "RESPONSABLE", "EQUIPO", "VERIFICADO", "ARCHIVO"])
    for m in db().execute("SELECT * FROM movimientos ORDER BY dueno, fecha"):
        ws.append([m["fecha"], m["tag"], m["patente"], m["concepto"], m["monto"],
                   m["dueno"], m["equipo"], "SI" if m["verificado"] else "NO", m["archivo"]])
    ws2 = wb.create_sheet("Gastos tarjetas")
    ws2.append(["FECHA", "CONCEPTO", "CATEGORIA", "MONTO", "BANCO", "RESUMEN"])
    for m in db().execute("""SELECT g.*, r.banco, r.archivo FROM gastos g
        JOIN resumenes r ON r.id=g.resumen_id ORDER BY r.id, g.id"""):
        ws2.append([m["fecha"], m["concepto"], m["categoria"], m["monto"], m["banco"], m["archivo"]])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"reporte_maestro_{datetime.now():%Y%m%d}.xlsx")

# ================= 5) CONFIGURACION =================
@app.route("/config")
def config():
    d = db()
    usuarios = d.execute("SELECT id,email,nombre,rol FROM usuarios ORDER BY id").fetchall()
    stats = {
        "movs": d.execute("SELECT COUNT(*) n FROM movimientos").fetchone()["n"],
        "gastos": d.execute("SELECT COUNT(*) n FROM gastos").fetchone()["n"],
        "flota": d.execute("SELECT COUNT(*) n FROM flota").fetchone()["n"],
        "mb": round(os.path.getsize(DB) / 1048576, 2) if os.path.exists(DB) else 0,
    }
    return render_template("config.html", usuarios=usuarios, stats=stats)

@app.route("/config/clave", methods=["POST"])
def cambiar_clave():
    nueva = request.form["nueva"].strip()
    if len(nueva) < 6:
        flash("La clave debe tener al menos 6 caracteres.")
    else:
        db().execute("UPDATE usuarios SET clave=? WHERE id=?",
                     (generate_password_hash(nueva), session["uid"]))
        db().commit(); flash("Clave actualizada.")
    return redirect("/config")

@app.route("/config/usuario", methods=["POST"])
def crear_usuario():
    if session.get("rol") != "admin":
        flash("Solo un administrador puede crear usuarios."); return redirect("/config")
    email = request.form["email"].strip().lower()
    clave = request.form["clave"].strip()
    nombre = request.form["nombre"].strip()
    if len(clave) < 6:
        flash("La clave debe tener al menos 6 caracteres.")
    else:
        try:
            db().execute("INSERT INTO usuarios(email,clave,nombre) VALUES(?,?,?)",
                         (email, generate_password_hash(clave), nombre))
            db().commit(); flash(f"Usuario {email} creado.")
        except sqlite3.IntegrityError:
            flash("Ese correo ya existe.")
    return redirect("/config")

@app.route("/config/usuario/borrar/<int:uid>", methods=["POST"])
def borrar_usuario(uid):
    if session.get("rol") != "admin":
        flash("Solo un administrador puede eliminar usuarios."); return redirect("/config")
    if uid == session["uid"]:
        flash("No podes eliminar tu propio usuario.")
    else:
        db().execute("DELETE FROM usuarios WHERE id=?", (uid,)); db().commit()
        flash("Usuario eliminado.")
    return redirect("/config")

init_db()
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=False)
